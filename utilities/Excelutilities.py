import openpyxl

def getRowCount(file,SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return (sheet.max_row)

def getColCount(file,SheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return  (sheet.max_column)

def readData(file,SheetName,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    return sheet.cell(row=rownum,column=colnum).value

def writeData(file,SheetName, rownum,colnum,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[SheetName]
    sheet.cell(rows=rownum,cols=colnum).value=data
    workbook.save(file)